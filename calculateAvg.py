import os
import openpyxl

xlsx_file = r'E:\joeCloud\Documents\工业互联网\工业互联网专项资金项目申报\审查' \
             r'\初审得分汇总表-基础支撑平台-中间工作表-程序计算.xlsx'

wb = openpyxl.load_workbook(xlsx_file)

sheet_hz = wb.get_sheet_by_name('基础平台汇总')

for row in range(4, 14):
    for col in range(5, 16):
        score_list = []
        for i in range(1,8):
            sheet_name = 'Sheet' + str(i)
            sheet_zj = wb.get_sheet_by_name(sheet_name)
            score = sheet_zj.cell(row=row, column=col).value
            if score is not None:
                score_list.append(score)
            else:
                score_list.append(0)
            score_list.sort(reverse=True)
        #print(score_list)
        avg = 0
        # 计算5个数的平均傎
        for j in range(1, 6):
            avg = score_list[j] + avg
        avg = avg/5
        #print(avg)
        sheet_hz.cell(row=row, column=col).value = avg

wb.save(xlsx_file)
##################################行业应用平台###############################
xlsx_file = r'E:\joeCloud\Documents\工业互联网\工业互联网专项资金项目申报\审查' \
             r'\初审得分汇总表-行业应用平台-中间工作表-程序计算.xlsx'

wb = openpyxl.load_workbook(xlsx_file)

sheet_hz = wb.get_sheet_by_name('行业平台汇总表')

for row in range(4, 17):
    for col in range(5, 17):
        score_list = []
        for i in range(1,8):
            sheet_name = 'Sheet' + str(i)
            sheet_zj = wb.get_sheet_by_name(sheet_name)
            score = sheet_zj.cell(row=row, column=col).value
            if score is not None:
                score_list.append(score)
            else:
                score_list.append(0)
            score_list.sort(reverse=True)
        #print(score_list)
        avg = 0
        # 去掉最高分和最低分，计算5个数的平均傎
        for j in range(1, 6):
            avg = score_list[j] + avg
        avg = avg/5
        #print(avg)
        sheet_hz.cell(row=row, column=col).value = avg

wb.save(xlsx_file)








