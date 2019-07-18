import os
import openpyxl
from docx import Document

doc_name =  os.getcwd()+'\\' + '形式审查(新版)-1.docx'
xlsx_file = os.getcwd()+'\\' + 'project-new.xlsx'
wb = openpyxl.load_workbook(xlsx_file)
#sheet = wb.get_sheet_by_name('基础支撑平台')
sheet = wb.get_sheet_by_name('行业应用平台')
doc = Document(doc_name)
tables = doc.tables
#print(len(tables))
table = tables[0]
item = {}
for i in range(6, 7):
    for j in range(1, 60):
        print('col=', j)
        print('value=',sheet.cell(row=i, column=j).value)

for i in range(6,  22):
    #取序号
    item['Num'] = sheet.cell(row=i, column=1).value
    #取项目名称
    item['name'] = sheet.cell(row=i, column=2).value
    table.cell(1, 3).text = item['name']
    #取单位名称
    item['company'] = sheet.cell(row=i, column=3).value
    table.cell(0, 3).text = item['company']
    #社会信用代码
    item['credNo'] = sheet.cell(row=i, column=12).value
    table.cell(0, 8).text = item['credNo']
    #项目总投资
    item['invest'] = sheet.cell(row=i, column=45).value
    table.cell(1, 8).text = str(item['invest'])
    #申报单位注册地
    item['address'] = sheet.cell(row=i, column=13).value
    table.cell(2, 3).text = item['address']
    # 联合申报单位名称
    item['allcompany'] = sheet.cell(row=i, column=6).value
    if item['allcompany'] is not None:
        table.cell(6, 2).text = item['allcompany']
    else:
        table.cell(6, 2).text = ''
    #取项目建设情况
    item['situation'] = sheet.cell(row=i, column=46).value
    table.cell(8, 3).text = item['situation']
    #取上线时间
    item['time'] = sheet.cell(row=i, column=44).value
    table.cell(8, 8).text = str(item['time'])[0:10]
    # 落地地点
    item['landaddress'] = sheet.cell(row=i, column=59).value
    table.cell(9, 3).text = str(item['landaddress'])

    #doc.save('A'+str(item['Num'])+'-'+item['name'] + '.docx')
    doc.save('B'+ str(item['Num']) + '-' + item['name'] + '.docx')

wb.save(xlsx_file)



