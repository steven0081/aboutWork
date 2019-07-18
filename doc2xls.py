import os
import openpyxl
import docx
from docx import Document

xlsx_file = os.getcwd() + '\\ok\\形式审查汇总表.xlsx'
# print(xlsx_file)
wb = openpyxl.load_workbook(xlsx_file)
# Sheet1 为基础支撑平台   Sheet2为行业应用平台
print('******************开始Word文件********************')
sheet_list = {'Sheet1': 'A', 'Sheet2': 'B'}
for key, value in sheet_list.items():
    print(key)
    sheet = wb.get_sheet_by_name(key)
    for i in range(2, sheet.max_row + 1):
        try:
            # A 为基础支撑平台   B 为行业应用平台
            doc_file = os.getcwd() + '\\ok\\' + value + str(sheet.cell(row=i, column=1).value) + '-' + str(
                sheet.cell(row=i, column=4).value).strip() + '-ok.docx'
            doc = Document(doc_file)
            # print(doc_file)
            tables = doc.tables
            # print(tables)
            table_1 = tables[0]
            table_2 = tables[1]
            # 社会信用代码
            sheet.cell(row=i, column=3).value = table_1.cell(0, 8).text
            # 总投资
            sheet.cell(row=i, column=5).value = table_1.cell(1, 8).text
            # 注册地
            sheet.cell(row=i, column=6).value = table_1.cell(2, 3).text
            # 是否符合申报主体要求
            sheet.cell(row=i, column=7).value = table_1.cell(2, 9).text
            # 所属行业
            sheet.cell(row=i, column=8).value = table_1.cell(3, 4).text
            # 是否符合申报主体行业要求
            sheet.cell(row=i, column=9).value = table_1.cell(3, 9).text
            # 是否资源池服务商
            sheet.cell(row=i, column=10).value = table_1.cell(4, 4).text
            # 服务商类型
            sheet.cell(row=i, column=11).value = table_1.cell(4, 9).text
            # 项目建设情况
            sheet.cell(row=i, column=12).value = table_1.cell(8, 3).text
            # 预计上线时间
            sheet.cell(row=i, column=13).value = table_1.cell(8, 8).text
            # 落地地点
            sheet.cell(row=i, column=14).value = table_1.cell(9, 4).text
            # 项目申报书,建设方案 财务报表 审计报告 两化融合评估报告 项目承诺书
            for a, b in zip(range(15, 21), range(11, 17)):
                content = table_1.cell(b, 6).text
                if content is not None:
                    sheet.cell(row=i, column=a).value = str(content)
            # 审查结论
            sheet.cell(row=i, column=21).value = table_1.cell(22, 2).text
            ##################取评分#####################
            for j, k in zip(range(22, 36), range(0, 18)):
                score = table_2.cell(k, 7).text
                # print(score)
                if score is not None:
                    sheet.cell(row=i, column=j).value = score
            # 总评分
            sheet.cell(row=i, column=36).value = table_2.cell(15, 3).text

        except Exception as e:
            print(e)

wb.save(xlsx_file)
